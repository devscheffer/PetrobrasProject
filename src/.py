import csv
import requests
from openpyxl import load_workbook
date_download_lst = [{"start":"2022-08-07","end":"2022-08-13"}]

for i in date_download_lst:
    file_url = f'https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/arquivos-lpc/resumo-semanal-lpc-{i.get("start")}-{i.get("end")}.xlsx'

    r = requests.get(file_url, stream = True)

    file_name = f'src/dataset/petrobras-{i.get("start")}-{i.get("end")}.xlsx'

    with open(file_name,"wb") as f1:
        for chunk in r.iter_content(chunk_size=1024):
            if chunk:
                f1.write(chunk)


    wb = load_workbook(file_name)
    sheets = wb.sheetnames
    for j in sheets:
        ws=wb[j]
        with open(f'src/dataset/{j}-{i.get("start")}-{i.get("end")}.csv', 'w',encoding='utf-8',newline="") as f:
            c = csv.writer(f)
            for r in ws.iter_rows():
                row_cells = []
                for cell in r:
                    data = cell.value
                    if cell.row < 10:
                        continue
                    row_cells.append(cell.value)
                if len(row_cells) > 0:
                    c.writerow(row_cells)

