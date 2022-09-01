
import csv
from openpyxl import load_workbook



def break_xls_to_csv(*,file_full_path,date,output_dir):
    date_start =date.get("start")
    date_end =date.get("end")
    wb = load_workbook(file_full_path)
    sheets = filter(lambda x: x in ["MUNICIPIOS"],wb.sheetnames)
    csv_list = []
    for j in sheets:
        ws=wb[j]
        csv_file_path =f'{output_dir}/{j}-{date_start}-{date_end}.csv'
        csv_list.append({"path":csv_file_path,"name":j,"date":{"start":date_start, "end":date_end}})
        with open(csv_file_path, 'w',encoding='utf-8',newline="") as f:
            c = csv.writer(f)
            for r in ws.iter_rows():
                row_cells = []
                for cell in r:
                    if cell.row < 10:
                        continue
                    row_cells.append(cell.value)
                if len(row_cells) > 0:
                    c.writerow(row_cells)
    return csv_list
