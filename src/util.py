
import csv
import requests
from openpyxl import load_workbook

def download_xls(*,output_dir,file_name,date):
    date_start =date.get("start")
    date_end =date.get("end")
    file_full_path = f'{output_dir}/{file_name}-{date_start}-{date_end}.xlsx'

    file_url = f'https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/precos/arquivos-lpc/resumo-semanal-lpc-{date_start}-{date_end}.xlsx'

    r = requests.get(file_url, stream = True)


    with open(file_full_path,"wb") as f1:
        for chunk in r.iter_content(chunk_size=1024):
            if chunk:
                f1.write(chunk)
    return file_full_path

def generate_csv_file(file_full_path,date,output_dir):
    date_start =date.get("start")
    date_end =date.get("end")
    wb = load_workbook(file_full_path)
    sheets = wb.sheetnames
    for j in sheets:
        ws=wb[j]
        with open(f'{output_dir}/{j}-{date_start}-{date_end}.csv', 'w',encoding='utf-8',newline="") as f:
            c = csv.writer(f)
            for r in ws.iter_rows():
                row_cells = []
                for cell in r:
                    if cell.row < 10:
                        continue
                    row_cells.append(cell.value)
                if len(row_cells) > 0:
                    c.writerow(row_cells)
